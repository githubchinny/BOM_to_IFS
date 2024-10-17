# %% [markdown]
# # BOM to IFS file
# 
# This script reads the latest collated BOM file from the sharepoint directory and extracts the columns we've configured for IFS migration.  This is a stop-gap while we are still working with excel files and will be replaced by a direct migration of data from 3DX straight into IFS - using IFS connect to drop files for processing on an internal IFS queue.
# 
# 
# Requirements:    
# A collated bom csv file in Engineering BoM sharepoint directory for the selected project, eg 'T33-BoM-XP_collated_BOM.csv'
# 
# 
# Inputs:
# Project name (T50, T50s, T33_XP...)
# 
# Outputs:   
# Writes compare files, Delta files, and migration txt files to:   
# 
# Output dir = sharepoint dir, project, IFS   
# 

# %%
import pandas as pd
import numpy as np
import os
import re
import io
import openpyxl
import excel_formatting
import logging
import argparse
import configparser

# %%
def type_of_script():
    '''
        determine where this script is running
        return either jupyter, ipython, terminal
    '''
    try:
        ipy_str = str(type(get_ipython()))
        if 'zmqshell' in ipy_str:
            return 'jupyter'
        if 'terminal' in ipy_str:
            return 'ipython'
    except:
        return 'terminal'

# %%
def update_parent_part(BOM):
    # reset index before trying to update, otherwise multiple rows get updated
    BOM.reset_index(inplace=True, drop=True)

    for sgroup, frame in BOM[BOM['Part Level'] > 2].groupby('Sub Group'):
        level = {}

        previous_parent_part=0

        for i, row in frame.iterrows():
            current_part_number = row['Part Number']
            current_part_level = row['Part Level']
            # reset higher levels for each assembly
            if current_part_level == 5:
                # remove entries from higher levels
                keys = [k for k in level if k > 5]
                for x in keys:
                    del level[x]

            # write part number to dictionary under current part level
            level[current_part_level] = current_part_number
            # update the current_parent_part if we have current part details (info from catia)
            # as we've created level 1 and 2 we don't need this check
            # print (current_part_level, current_part_number)
            level[2] = group_area_dict[sgroup]
            if i > 0:
            # get the max part level from the level dictionary that's less than current part level
                previous_parent_level = max(k for k in level if k < current_part_level)

                    # update the parent part
                # print (i, "Parent part {} from previous level {}".format(level[previous_parent_level], previous_parent_level))
                BOM.at[i,'Parent Part'] = level[previous_parent_level]
    return BOM

# %%
def db_pool_connection():
    # connection to database using sqlalchemy
    import oracledb
    from sqlalchemy import create_engine
    from sqlalchemy import text
    import db_config
    import pandas as pd

    # d = r"C:\Users\mark.chinnock\oracle\instantclient_21_10"
    # oracledb.init_oracle_client(lib_dir=d)

    pool = oracledb.create_pool(user=db_config.user, password=db_config.LIVE_userpwd, dsn=db_config.LIVE_connect_string,
                                min=1, max=5, increment=1)

    return pool

# %%
def get_ifs_part_cat(env):
    # connection to database using sqlalchemy
    import oracledb
    from sqlalchemy import create_engine
    from sqlalchemy import text
    import db_config
    import pandas as pd
    from contextlib import suppress

    # Database Credentials
    username = db_config.user
    password = db_config.LIVE_userpwd

    engine = create_engine(
        f'oracle+oracledb://:@',
            thick_mode=None,
            connect_args={
                "user": db_config.user,
                "password": db_config.LIVE_userpwd,
                "host": db_config.LIVE_host,
                "port": 1521,
                "service_name": db_config.LIVE_service
        })

    query = ("select distinct c.part_no, c.unit_meas, p.lot_tracking_code, p.serial_tracking_code, p.serial_rule "
            "from ifsapp.inventory_part c "
            "left join ifsapp.part_catalog p "
            "on c.part_no = p.part_no "
            "where c.part_no like 'T%'")

    try:
        with engine.connect() as connection:
            # print(connection.scalar(text("""SELECT * from IFSAPP.purchase_order_line_all""")))
            query = connection.execute(text(query))

        df = pd.DataFrame(query.fetchall())

        df.columns = df.columns.str.upper()


    except:
        pass

    

    return df


# %%
if type_of_script() == 'terminal':
    parser = argparse.ArgumentParser()
    parser.add_argument("project", metavar='Project', type=str, help="T50, T50s, T33_XP, etc") 
    parser.add_argument("timestamps", metavar='prev_timestamps', type=str, help='timestamp portion from previous migration filename: eg 20230530-2044')
    parser.add_argument("env", metavar='Environment', type=str, help='LIVE or Sandbox')
    # parser.add_argument("incrementer", metavar='Incrementer', type=int)    
    parser.add_argument('--disable-delta', help='Use this to switch delta file processing off', action='store_true')
    parser.add_argument('-i', '--ignore', action='append', help="Cols to Ignore from comparison - only needed when we've added/removed a column from migration files")

    args = parser.parse_args()
    project = args.project
    prev_timestamps = args.timestamps
    env=args.env
    DELTA = not args.disable_delta
    ignore_cols_for_comparison = args.ignore
    # incrementer=args.incrementer

else:
    # set defaults if we're running in jupyter
    project = 'T50'
    # for producing the delta files
    prev_timestamps = '20231205-1618'
    env = 'LIVE'
    ignore_cols_for_comparison = None
    # incrementer=6
    # env = 'Sandbox'

print ("ignore cols {}".format(ignore_cols_for_comparison))

# personal one drive
user_dir = 'C:/Users/USERNAME'
# replace USERNAME with current logged on user
user_dir = user_dir.replace('USERNAME', os.getlogin())

pattern = re.compile(r"-|_")
project_uc = pattern.split(project.upper())[0]

# go find the Engineering BoM directory within the User directory
# from glob import glob
# sharepoint_dir = glob(user_dir + "/*gordonmurraydesign*/Documents - Engineering BoM*", recursive = True)[0]

# read in config file
config = configparser.ConfigParser()
config_file = user_dir + '/user_directory.ini'
config.read(config_file)

# read in gm_dir and gm_docs from config file
try:
    gm_dir = config[os.getlogin().lower()]['gm_dir']
    gm_docs = config[os.getlogin().lower()]['gmd']
except:
    print ("You're probable missing a valid {} file".format(config_file))
    SystemExit(1)

# this may find more than one sharepoint directory
sharepoint_dir = user_dir + "/" + gm_dir + "/" + gm_docs

base = os.path.join(sharepoint_dir, project)
# base = user_dir

# where we'll write out the files
outdir = os.path.join(base, 'IFS')

import time
timestr = time.strftime("%Y%m%d-%H%M")

# logfile name and location
logfile = os.path.join(sharepoint_dir, project, 'IFS', 'logs', "BOM_to_IFS_{}_{}_log.txt".format(project, timestr))
logging.getLogger('matplotlib').setLevel(logging.WARNING)
logging.basicConfig(filename=logfile, filemode='a', level=logging.DEBUG, format='%(asctime)s %(levelname)s:%(message)s')

logit = logging.getLogger(__name__)

logit.info("Starting the process...")
logit.info("Creating files for project {}".format(project))
logit.info("Running from {}".format(type_of_script()))
logit.info("base: {}".format(base))
logit.info("outdir: {}".format(outdir))
logit.info("Columns being ignore in comparison later (new cols?): {}".format(ignore_cols_for_comparison))

# %%
group_area_dict = {
'A02-Panels & Closure Systems':project +'-01', 
'A03-Exterior Systems':project +'-01',
'A01-Structure Systems':project +'-01', 
'B01-Suspension Systems':project +'-02',
'C01-Braking Systems':project +'-02', 
'D01-Steering Systems':project +'-02', 
'E01-Pedal System':project +'-02',
'M01-Control Systems':project +'-03', 
'M02-Traction Systems':project +'-03',
'M03-Electrical Distribution Sys':project +'-03', 
'M04-Multimedia Systems':project +'-03',
'M05-Safety & Security Systems':project +'-03', 
'M06-Software Systems':project +'-03',
'N01-Interior & Trim Systems':project +'-04', 
'N02-HVAC Systems':project +'-04',
'F01-ICE Powertrain Systems':project +'-03', 
'G01-Transmission Systems':project +'-03',
'J01-Pwt NVH & Heatshield Sys':project +'-03', 
'L01-Cooling Systems':project +'-03',
'R01-Styling':project +'-07',
'P01-Packaging':project +'-06',
'T01-Tooling':project +'-08',
'U01-Development':project +'-09',
'V01-Accessories':project +'-10'
}

# %%
part_type_dict = {
'AIH':'Manufactured',
'BOF':'Purchased',
'BOP':'Purchased',
'CON':'Purchased (Raw)',
'ENG':'Purchased (Raw)',
'FAS':'Purchased (Raw)',
'FIP':'Purchased (Raw)',
'RAW':'Purchased (Raw)',
'MIH':'Manufactured',
'POA':'Manufactured',
'MOB':'Manufactured'
}


# %%
# existing_bom_file = 'T33-BoM-XP_collated_BOM.csv'
existing_bom_file = '{}_collated_BOM.csv'.format(project)

path = os.path.join(base, existing_bom_file)
try:
    with open(path, "rb") as f:

        existing_bom = pd.read_csv(f, na_values='*', parse_dates=True, low_memory=False) 
        # sheetnames = [sheet for sheet in f.sheet_names]
except (FileNotFoundError):
    logit.critical("File not found: {}".format(path))



# %%
# call IFS for latest parts catalogue info

# get db pool connection
pool = db_pool_connection()

query = ("select distinct c.part_no, c.unit_meas, p.lot_tracking_code, p.serial_tracking_code, p.serial_rule "
        "from ifsapp.inventory_part c "
        "left join ifsapp.part_catalog p "
        "on c.part_no = p.part_no "
        "where c.part_no like 'T%'")

with pool.acquire() as connection:
    ifs_parts_cat = pd.read_sql(query, connection)



# ifs_parts_cat = get_ifs_part_cat(env)



# %%
# was using this to identify data with wrong REV_NO in IFS versus No of rows
# query = ("select * from ifsapp.eng_part_revision_reference where part_no like 'T%'")

# with pool.acquire() as connection:
#     test = pd.read_sql(query, connection)

# test2 = test.groupby('PART_NO').agg({'PART_NO':'count','REV_NO':'max'}).rename(columns={'PART_NO':'count','REV_NO':'max_rev'})
# test3 = test2[test2['count'] < test2['max_rev']].reset_index()
# test3[test3['PART_NO'].str.len() < 10]

# import xlwings as xw

# wb = xw.Book()
# ws = wb.sheets[0]

# ws['A1'].options(pd.DataFrame, header=1, index=True).value=test3

# %%
# parts_cat_dict for looking up serial tracking setting
parts_cat_dict = {}	
parts_cat_dict = pd.Series(ifs_parts_cat['LOT_TRACKING_CODE'].values,index=ifs_parts_cat['PART_NO']).to_dict()

ifs_parts_dict = {}
ifs_parts_dict = pd.Series(ifs_parts_cat['UNIT_MEAS'].values,index=ifs_parts_cat['PART_NO']).to_dict()    

ifs_serial_tracking = {}
ifs_serial_tracking = pd.Series(ifs_parts_cat['SERIAL_TRACKING_CODE'].values,index=ifs_parts_cat['PART_NO']).to_dict()


existing_bom.reset_index(inplace=True)
existing_bom.rename(columns={'index':'orig_sort'}, inplace=True)

# %%
# we're not expecting rows for Part level < 3 at the moment.  The assumption is our extracted Excel BoM starts at Part Level 3
existing_bom = existing_bom[existing_bom['Part Level'] > 2]

# %%
parts_reg_file = 'collated_parts_register.xlsx'
path = os.path.join(sharepoint_dir, parts_reg_file)
try:
    parts_reg = pd.read_excel(path)
except (FileNotFoundError):
    logit.critical("File not found: {}".format(path))

engineering_names_file = 'Engineer_names.xlsx'
path = os.path.join(sharepoint_dir, engineering_names_file)
try:
    engineering_names = pd.read_excel(path, sheet_name='Users')
except (FileNotFoundError):
    logit.critical("File not found: {}".format(path))

engineering_names = engineering_names[engineering_names['LAST_NAME'].isna()==False]

# %%
# parts_reg_dict for looking up Engineer
parts_reg_dict = {}
parts_reg_dict = pd.Series(parts_reg['Engineer'].values,index=parts_reg['Part Number']).to_dict()


# %%
# parts_reg_dict for looking up Engineer
eng_name_dict = {}
eng_name_dict = pd.Series(engineering_names['PERSON_ID'].values,index=engineering_names['LAST_NAME']).to_dict()

initials_dict = {}
initials_dict = pd.Series(engineering_names['PERSON_ID'].values,index=engineering_names['INITIALS']).to_dict()



# %%
# I think this will either work or fail quietly.
existing_bom.rename(columns={'Res. Des. Engineer':'Engineer'}, inplace=True)

# %%
# this part is a problem.  It's wrong in the BoM and hasn't been fixed.  It's a U part so won't make it onto a PROD car.
i = existing_bom[(existing_bom['Part Number'] == 'T50-U4395*') & (existing_bom['Issue Level'] == 1)].index
existing_bom.drop(i, inplace=True)

# %% [markdown]
# # Bring only required rows from BoM
# 
# This is only relevant for T50 as other BoMs don't have the same 'Requirement' columns.
# 
# Same logic as MBOM - bring only rows that don't have 4 Ns in the 4 Requirement cols
# 
# 
# |index|Part Number|	VP Requirement|	PP Requirement|	PS Requirement|	Prod Requirement|
# |-----|-----------|---------------|---------------|---------------|-----------------|
# |0|	T50-A02|	Y|	Y|	Y|	Y|
# |1|	T50-A02-01|	Y|	Y|	Y|	Y|
# |2|	T50-A0039|	Y|	Y|	Y|	Y|
# |3|	T50-A2889|	Y|	Y|	Y|	Y|
# |4|	T50-A0041|	Y|	Y|	Y|	Y|
# ...	...	...	...	...	...
# |33927|	TFF-AA040|	N|	N|	Y|	Y|
# |33928|	TPP-LZ065|	N|	N|	Y|	Y|
# |33929|	T50-L0381|	Y|	Y|	Y|	Y|
# |33930|	T50-L0373|	Y|	Y|	Y|	Y|
# |33931|	TFF-SA888|	Y|	Y|	Y|	Y|
# 
# 
# 
# convert the 'Y/N' to 1/0 and then sum them.

# %%
include_aftersales = False

if project == 'T50':

    # map 'N' to 0 and anything else to 1; then sum them up - keep > 0
    existing_bom['VP Requirement'] = np.where(existing_bom['VP Requirement'] == 'N', 0, 1)
    existing_bom['PP Requirement'] = np.where(existing_bom['PP Requirement'] == 'N', 0, 1)
    existing_bom['PS Requirement'] = np.where(existing_bom['PS Requirement'] == 'N', 0, 1)
    existing_bom['Prod Requirement'] = np.where(existing_bom['Prod Requirement'] == 'N', 0, 1)

    # make sure we have capitals
    existing_bom['Service Identifier'] = existing_bom['Service Identifier'].str.upper()
    existing_bom2 = pd.DataFrame()
    existing_bom2 = existing_bom[existing_bom.filter(regex='Requirement').sum(axis=1)>0]

    # include AFTERSALES build reference rows as well
    aftersales = existing_bom[( ( existing_bom['Build References'].str.contains('AFTERSALES', na=False) ) & ( existing_bom['Service Identifier'].isin(['Y','C']) ) ) ]
    
    # waiting before including these rows
    if include_aftersales:
        existing_bom2 = pd.concat([existing_bom2, aftersales])
        
    # make sure we haven't created duplicates
    existing_bom2.drop_duplicates(subset='orig_sort', inplace=True)
    # make sure we've retained the original order
    existing_bom2.sort_values('orig_sort', inplace=True)

else:
    # pass through existing_bom for all other projects
    existing_bom2 = existing_bom



# %%
try:
    existing_bom2.loc[:,'Part Level'] = existing_bom2['Part Level'].replace({r'\*': ''}, regex=True)
    existing_bom2.loc[:,'Part Number'] = existing_bom2['Part Number'].replace({r'\*': ''}, regex=True)
    existing_bom2.loc[:,'Parent Part'] = existing_bom2['Parent Part'].replace({r'\*': ''}, regex=True)
except (ValueError):
    logit.error("Can't replace asterisk for {}".format(existing_bom2[['Function Group','Sub Group','Part Number']]))
    print ("Can't replace asterisk for {}".format(existing_bom2[['Function Group','Sub Group','Part Number']]))

existing_bom2.loc[:,'Part Number'] = existing_bom2['Part Number'].str.upper()
existing_bom2.loc[:,'Parent Part'] = existing_bom2['Parent Part'].str.upper()

# %%
# keeping Engineer data to show what it comes through as
cols = ['orig_sort',
'Function Group',
'Sub Group', 
'Part Level', 
'Part Number',
'Issue Level',
'Part Description', 
'Part - Qty',
'Parent Part', 
'Source Code',
'Weight (KG)',
'Release Status',
'Engineer',
'Service Identifier']

# merged = pd.merge(df, existing_bom[cols], on='Part Number', how='left', indicator=True)
IFS = existing_bom2[cols]

# %%
# drop Packaging function group as only WIP info.  
logit.info("Not processing Packaging Function Group")
IFS = IFS[IFS['Function Group'] != 'Packaging']

# ignore Accessories for the time being as it is not structured correctly (no level 4)
# 2023-05-11 Matt Perry says we need to include these and has corrected the level 4 issue
# IFS = IFS[IFS['Function Group'] != 'Accessories']

# %%
# # # update Engineer to what we get back from parts_reg.  This is step 1
# although this gets set we will ignore it later as the data isn't good enough
IFS['Engineer lookup'] = IFS['Part Number'].map(parts_reg_dict).fillna('NO_ENGINEER')

# # # next, lookup the surname in the engineer dataframe to get the PERSONID


# %%
# IFS[IFS['Part Number'] == IFS['Parent Part']]
# print (IFS.shape)
IFS.dropna(subset='Part Number', inplace=True)
# print (IFS.shape)

# %%
# source_codes_df = pd.DataFrame(data=source_codes_dict, index=source_codes_dict.keys())
IFS['Part Type'] = IFS['Source Code'].map(part_type_dict)


# %% [markdown]
# # Empty Structures
# 
# Lorena says if empty structure for 'BOF,'BOP', then it should be purchased (Raw) rather than purchased
# 
# Empty structure = following row has part level <= current part level
# 
# *** Even though this logic works, we don't supply Part Type to the migration process and Migration itself does this empty structure logic instead ***

# %%
# if BOF or BOP and next row's part level <= current part level update to Purchased (Raw)
#shift(-1) reads the next row
IFS['Part Type'] = np.where(IFS['Source Code'].isin(['BOF','BOP']) & (IFS['Part Level'].shift(-1) <= IFS['Part Level']), 'Purchased (Raw)', IFS['Part Type'])


# %% [markdown]
# # Drop Source Codes
# 
# 03/03/2023: we are going to drop POA rows completely from the files and then correct the parent part to maintain the structure   
# 08/03/2023: we are dropping ENG rows completely from the dataframe and correct the parent part, if needed   
# 09/03/2023: we are dropping SOP rows as well now   
# 21/07/2023: we want to include SOP rows where there is a Service Identifier of 'Y' or 'C' and it is configured for a PROD vehicle
# 

# %%
# drop source codes
# IFS = IFS[~IFS['Source Code'].isin(['POA','ENG'])]
IFS = IFS[~IFS['Source Code'].isin(['POA','ENG','SOP'])]

# %%
pd.crosstab(IFS['Service Identifier'], IFS['Source Code'])


# %%
IFS[['Source Code']].value_counts(dropna=False)


# %% [markdown]
# # Level 1 and 2 rows
# Create a Level 1 and 2 for each function group as they don't exist in the BoM

# %%
# this is clearly wrong but can't work out whether it's worth fixing for IFS.  Needs some proper thought around which rows we add up, and when
car_weight = IFS[IFS['Part Level'] >= 5]['Weight (KG)'].sum()

level1 = {'Function Group': None,
 'Sub Group':None,
 'Part Level':1,
 'Part Number':[project + '-CAR'],
 'Issue Level':1,
 'Part Description': ['CAR'],
 'Part - Qty':1,
 'Parent Part':np.NaN,
 'Source Code':'SYS',
 'Weight (KG)':car_weight,
 'Release Status':'REL',
 'Engineer':['Engineer Name']
 }

level2 = {'Function Group': None,
 'Sub Group':['BODY SYSTEMS','CHASSIS SYSTEMS','POWERTRAIN SYSTEMS','ELECTRICAL SYSTEMS','INTERIOR & HVAC SYSTEMS','PACKAGING','STYLING','TOOLING','DEVELOPMENT','ACCESSORIES'],
 'Part Level':2,
 'Part Number':[project + '-01',project + '-02',project + '-03',project + '-04',project + '-05',project + '-06',project + '-07',project + '-08',project + '-09',project + '-10'],
 'Issue Level':1,
 'Part Description': ['BODY SYSTEMS','CHASSIS SYSTEMS','POWERTRAIN SYSTEMS','ELECTRICAL SYSTEMS','INTERIOR & HVAC SYSTEMS','PACKAGING','STYLING','TOOLING','DEVELOPMENT','ACCESSORIES'],
 'Part - Qty':1,
 'Parent Part':project + '-CAR',
 'Source Code':'SYS',
 'Weight (KG)':1000,
 'Release Status':'REL',
 'Engineer':['Engineer Name','Engineer Name','Engineer Name','Engineer Name','Engineer Name','Engineer Name','Engineer Name','Engineer Name','Engineer Name','Engineer Name']
 } 

level1_df = pd.DataFrame(level1)
level2_df = pd.DataFrame(level2)

level2_df

# use the order of concat to put l1 at top, then l2, and then the rest of IFS 
IFS = pd.concat([level1_df, level2_df, IFS])

# %%
IFS.reset_index(inplace=True, drop=True)


# %% [markdown]
# ## Sum function group weights
# groupby the function groups, update the corresponding level 2 with the sum of that group.

# %%
function_groups_dict = {
    project + '-01':['A'], 
    project + '-02':['B','C','D','E'], 
    project + '-03':['F','G','J','L'],
    project + '-04':['M'],
    project + '-05':['N'],
    project + '-06':['P'],
    # 'T50-07':['R'],
    project + '-08':['T'],
    # 'T50-09':['U']
    }

sum_function_groups = IFS.groupby(IFS['Sub Group'].str[:1])['Weight (KG)'].sum()

for key in function_groups_dict:
    IFS.loc[:,'Weight (KG)'] = np.where(IFS['Part Number'] == key, sum_function_groups[function_groups_dict[key]].sum(), IFS['Weight (KG)'])

# key = 'T50-01'



# %%
# this is a better car weight sum
car_sum = 0
for key in function_groups_dict:
    car_sum = car_sum + IFS['Weight (KG)'][IFS['Part Number'] == key].values
    # print (IFS['Weight (KG)'][IFS['Part Number'] == key].values)

IFS['Weight (KG)'][IFS['Part Number'] == 'T50-CAR'] = car_sum


# %% [markdown]
# # Cleanse Dataframe

# %%
def cleanse_dataframe(df):

    df = df.replace(r'^\W', '', regex=True)
    # remove any number of leading or trailing spaces
    df = df.replace(r'^ +| +$', '', regex=True)
    # remove strange values at the end of part descriptions
    df = df.replace(r'_x000D_', '', regex=True)
    # remove any remaining \n characters
    df = df.replace(r'\n', '', regex=True)

    try:
        # do individually as might put asterisk in any of the columns
        # df = df.replace(regex='\\*', value='')
        df['Part Level'] = df['Part Level'].replace({r'\*': ''}, regex=True)
        df['Part Number'] = df['Part Number'].replace({r'\*': ''}, regex=True)
        df['Parent Part'] = df['Parent Part'].replace({r'\*': ''}, regex=True)
    except (ValueError):
        print ("Can't replace asterisk for {}".format(df[['Function Group','Sub Group','Part Number']]))
        logit.error("Can't replace asterisk for {}".format(df[['Function Group','Sub Group','Part Number']]))

    # leaving these as a reminder that they don't work!  str.replace \* doesn't actually remove the asterisk! 
    # dict_df[sheet]['Parent Part'] = dict_df[sheet]['Parent Part'].str.replace(r'\s*', '', regex=True)
    # dict_df[sheet]['Part Number'] = dict_df[sheet]['Part Number'].str.replace(r'\s*', '', regex=True)
    # .replace is better at removing the asterisks
    
    #uppercase Release Status
    df['Release Status'] = df['Release Status'].str.upper()

    return df

# %%
# remove any whitespace/carriage return characters
cleansed_df = cleanse_dataframe(IFS)

# %% [markdown]
# # Parent Part
# Correct the parent part before writing it out.  After removing the POA rows we will need to correct the parent part reference for child parts

# %%
# call update_parent_part which updates the bom directly 
cleansed_df = update_parent_part(cleansed_df)

# %% [markdown]
# # Build an SA Index
# 
# The same parts can be used all over the car/bom.  We need a way of identifying the parts that fall below a specific assembly so that we can collate like parts together and calculate the required quantities.  
# 
# 1. I'm looking for Level 4 or 5s and setting the SA_Index to be the unique row number that came from the 3dx extract.   
# 2. Less than level 4 get their unique row number 
# 3. This leaves > level 5 to forward fill with whatever the level 5 row number above was.  This becomes the base for the SA_Index
# 4. Concatenate the individual level and part number (Title) onto the SA_Index

# %%
cleansed_df['SA_Index'] = np.where(cleansed_df['Part Level'].isin([4,5]), cleansed_df['orig_sort'].astype(str), np.nan)
cleansed_df['SA_Index'] = np.where(cleansed_df['Part Level'] < 4, cleansed_df['orig_sort'].astype(str), cleansed_df['SA_Index'])
# forward fill so that > Level 5 get the same index
cleansed_df['SA_Index'] = cleansed_df['SA_Index'].ffill()
# don't include Part level in SA_Index
# cleansed_df['SA_Index'] = cleansed_df['SA_Index'].astype(str) + '_' + cleansed_df['Part Level'].astype(str) + '_' + cleansed_df['Part Number']
cleansed_df['SA_Index'] = cleansed_df['SA_Index'].astype(str) + '_' + cleansed_df['Part Number']

# %% [markdown]
# # Correct multiple issue levels
# IFS migration cannot handle a part number having more than one revision level in the same migration file.  Need to go through and find the latest Issue Level for each part and update all instances to that
# 
# Might need to do this later when we have an SA_Index to drop the whole assembly, where applicable

# %%
# group by pn and issue level, unstack to make issue levels columns
unstacked = cleansed_df.groupby(['Part Number','Issue Level']).size().unstack()

# find number of columns dynamically, as number of unique status controls the number of columns
expected_status_count = len(unstacked.columns) - 1

unstacked2 = unstacked[unstacked.isna().sum(axis=1)!=expected_status_count].reset_index()
dup_parts = unstacked2['Part Number'].tolist()

# find level 5 rows (assy) that are in the dup_parts list
dup_assy = cleansed_df[(cleansed_df['Part Number'].isin(dup_parts)) & (cleansed_df['Part Level'] ==5)].sort_values(by=['Part Number','Issue Level'])
# drop duplicates on pn and keep the last knowing we've sorted them so last row will be max issue level
max_assy = dup_assy.drop_duplicates(subset=['Part Number'], keep='last')

# substract max_assy from dup_assy, which will leave all the non max rows.
SA_to_remove = set(dup_assy['SA_Index'].str.split('_').str[0].drop(max_assy.index).tolist())

# create dataframe to write to
to_delete = pd.DataFrame()

#loop through SA_to_remove and look for the part SA_Index in SA_Index
for x in SA_to_remove:
    temp_df = cleansed_df[cleansed_df['SA_Index'].str.split('_').str[0] == (x)]
    to_delete = pd.concat([to_delete, temp_df])

# drop the to_delete datatframe matches on index from cleansed_df
cleansed_df.drop(to_delete.index, inplace=True)

# %%
# the idea is any parts remaining with different revisions should just use the latest(max) revision
# group by the part number and make a dictionary of the max revision

max_issue_levels = cleansed_df.groupby(['Part Number'])['Issue Level'].max().reset_index()
# max_issue_level for looking up max_issue of a part in the T50 bom.  Shouldn't be a problem in T33
max_issue_dict = {}	
max_issue_dict = pd.Series(max_issue_levels['Issue Level'].values,index=max_issue_levels['Part Number']).to_dict()

cleansed_df['Issue Level'] = cleansed_df['Part Number'].map(max_issue_dict)

# %%
to_delete.shape

# %% [markdown]
# # Correct FIPs 
# 
# FIP parent part needs to be the previous BOF
# 

# %%
cleansed_df['BOF_Parent'] = np.where(cleansed_df['Source Code'] == 'BOF', cleansed_df['Part Number'], np.nan)
cleansed_df['BOF_Parent'].ffill(inplace=True)
cleansed_df['Parent Part'] = np.where(cleansed_df['Source Code'] == 'FIP', cleansed_df['BOF_Parent'], cleansed_df['Parent Part'])

# %% [markdown]
# # Group by Duplicate Parts within SA 
# 
# Where we have repeated a part / parent part within an assy, we only need to tell IFS about the structure once and sum up the quantities into one row.
# 
# Need to make sure the description for a part number is unique throughout the file, regardless of issue level.

# %%
# sort out quantity and drop duplicates

cleansed_df['Quantity'] = cleansed_df.groupby(['SA_Index','Part Number','Parent Part','Issue Level'])['Part - Qty'].transform('sum')

cleansed_df.drop(columns='Part - Qty', inplace=True)
cleansed_df.rename(columns={'Quantity':'Part - Qty'}, inplace=True)
cleansed_df.drop_duplicates(subset=['SA_Index','Part Number','Parent Part','Issue Level','Part - Qty'], inplace=True)



# %%
# there is a row with all NaN - remove it
cleansed_df.dropna(how='all', inplace=True)


# %% [markdown]
# # Create structure_df and parts_df

# %%
# only create structure_df with 4 cols, using Issue Level once - otherwise we can't rename the columns without it renaming both Issue Levels
# we'll add the PART_REV later by mapping to the parent_part_dict

structure_cols = [
'Parent Part',
'Issue Level',
'Part Number',
'Part - Qty',
]

# bring in Release Status from BoM as well
part_cols = [
'Part Number',
'Part Level',
'Part Description',
'Weight (KG)',
'Source Code',
'Issue Level',
'Function Group',
'Sub Group',
'Engineer lookup',
'Release Status'
]

# %%
# don't need to bring the level1 from cleansed_df into the structure file.  Level1 will always be the first row
structure_df = cleansed_df[1:][structure_cols].copy()
parts_df = cleansed_df[part_cols].copy()

# rename cols to match IFS naming
structure_df.rename(columns={
    'Parent Part':'PART_NO', 
    'Part Number':'SUB_PART_NO',
    'Part - Qty':'QTY',
    'Issue Level':'SUB_PART_REV'   
}, inplace=True)

parts_df.rename(columns={
    'Part Number':'PART_NO',
    'Part Description':'DESCRIPTION',
    'Weight (KG)':'WEIGHT_NET',
    'Function Group':'FUNCTION_GROUP',
    'Sub Group':'SUB_GROUP',
    'Part Level':'PART_LEVEL',
    'Engineer lookup':'PART_RESPONSIBLE',
    'Source Code':'SOURCE_CODE',
    'Release Status':'RELEASE_STATUS'}
    , inplace=True)


# %%
# find the issue level for the parent part.  Then map this to the PART_REV column of the structure_df

parent_part_dict = {}	
# parent_part_dict = pd.Series(IFS['Issue Level'][IFS['Part Number'].isna()==False].values,index=IFS['Part Number'][IFS['Part Number'].isna()==False]).to_dict()
# changed this to use cleansed_df
parent_part_dict = pd.Series(cleansed_df['Issue Level'][cleansed_df['Part Number'].isna()==False].values,index=cleansed_df['Part Number'][cleansed_df['Part Number'].isna()==False]).to_dict()

# this creates the PART_REV column
structure_df['PART_REV'] = structure_df['PART_NO'].map(parent_part_dict)

# now correct the column ordering for the final file template
structure_df = structure_df[['PART_NO', 'PART_REV', 'SUB_PART_NO', 'QTY', 'SUB_PART_REV']]



# %%
parts_df['INFO_TEXT'] = 'Drawing URL?'
# map LOT_TRACKING_CODE from existing parts, 22/06/2023 - changed to default to Not Lot Tracking.  This is usual default
parts_df['LOT_TRACKING_CODE'] = parts_df['PART_NO'].map(parts_cat_dict).fillna('Not Lot Tracking')
# parts_df['LOT_TRACKING_CODE'] = 'Order Based'
parts_df['SERIAL_RULE'] = 'Manual'
parts_df['CONFIGURABLE'] = 'Not Configured'
parts_df['AQUISITION_CODE'] = 'Demand'
parts_df['PLANNING_METHOD'] = 'Standard Planned'
# use Engineer name from BoM to demonstrate data quality
# parts_df['PART_RESPONSIBLE'] = 'Engineer Name'
# lookup part issue level for ENG_REV_NO.
parts_df['ENG_PART_REV'] = parts_df['Issue Level']
# lookup if this is a parent part with a different issue level, otherwise leave as the part issue level
parts_df['ENG_PART_REV'] = parts_df['PART_NO'].map(parent_part_dict).fillna(parts_df['ENG_PART_REV'])
# look up existing part information for serial tracking first
parts_df['SERIAL_TRACKING_CODE'] = parts_df['PART_NO'].map(ifs_serial_tracking)
# Use this rule for anything we didn't lookup - Serial Tracking for levels 1, 2 and 5, otherwise keep the current value
# parts_df['SERIAL_TRACKING_CODE'] = np.where((parts_df['SERIAL_TRACKING_CODE'].isna() & parts_df['PART_LEVEL'].isin([1,2,5])), 'Serial Tracking', parts_df['SERIAL_TRACKING_CODE'])
# and then set all remaining isna() to Not Serial Tracker
parts_df['SERIAL_TRACKING_CODE'] = np.where(parts_df['SERIAL_TRACKING_CODE'].isna(), 'Not Serial Tracking', parts_df['SERIAL_TRACKING_CODE'])
# map UNIT CODE to existing value, didn't leave any NaN
parts_df['UNIT_CODE'] = parts_df['PART_NO'].map(ifs_parts_dict)
# AIH, MIH, MOB, POA, SYS source codes are Make, everything else is Buy.  This should match with Manufactured in part_type
parts_df['PROVIDE'] = np.where(parts_df['SOURCE_CODE'].isin(['AIH','MIH','MOB','POA','SYS']),'Make','Buy')
# map inventory part planning of 'A' and 'P' using the make or buy decision above
parts_df['INVENTORY_PART_PLANNING'] = np.where(parts_df['PROVIDE']=='Make', 'P', 'A')
# part level 1 is exception to the rule above and will always be PROVIDE = 'Make', INVENTORY_PART_PLANNING = 'A'
parts_df['INVENTORY_PART_PLANNING'] = np.where(parts_df['PART_LEVEL'] == 1, 'A', parts_df['INVENTORY_PART_PLANNING'])
# create default status for inventory part status of 'A' for purchase/purchase raw, and 'I' for make parts
parts_df['INVENTORY_PART_STATUS'] = np.where(parts_df['PROVIDE']=='Make', 'A', 'I')

# don't need these columns anymore
parts_df.drop(columns=['Issue Level'], inplace=True)


# %%
# default weights to zero where not provided
parts_df['WEIGHT_NET'] = np.where(parts_df['WEIGHT_NET'].isna(), 0, parts_df['WEIGHT_NET'])

# PART_NO cannot contain lowercase 
parts_df['PART_NO'] = parts_df['PART_NO'].str.upper()

# get rid of negative weights whilst we're waiting for BoM to be corrected
parts_df['WEIGHT_NET'] = np.where(parts_df['WEIGHT_NET'] < 0, 0, parts_df['WEIGHT_NET'])

# add blank VARIANT and MATURITY cols for future
parts_df[['VARIANT','MATURITY']] = np.NaN


# %%
# ENG_REV_NO is an IFS internally incremented number.  We don't have control over it so we are not going to pass it.
# migration script in IFS will handle this

part_cols_ordered = ['PART_NO',
'DESCRIPTION',
'WEIGHT_NET',
'INFO_TEXT',
'UNIT_CODE',
'LOT_TRACKING_CODE',
'SERIAL_RULE',
'SERIAL_TRACKING_CODE',
'CONFIGURABLE',
'PROVIDE',
'AQUISITION_CODE',
'PLANNING_METHOD',
'PART_RESPONSIBLE',
'ENG_PART_REV',
# 'ENG_REV_NO',
'FUNCTION_GROUP',
'SUB_GROUP',
'PART_LEVEL',
'SOURCE_CODE',
'VARIANT',
'MATURITY',
'INVENTORY_PART_PLANNING',
'RELEASE_STATUS',
'INVENTORY_PART_STATUS'
]

parts_df = parts_df[part_cols_ordered]

# %%
# fill blank UNIT_CODE with LTR for FLA
parts_df['UNIT_CODE'] = np.where(((parts_df['UNIT_CODE'].isna()) & (parts_df['SOURCE_CODE'] == 'FLA')), 'LTR', parts_df['UNIT_CODE'])

# fill remaining blank UNIT_CODE with PCS
parts_df['UNIT_CODE'] = parts_df['UNIT_CODE'].fillna('PCS')


# %% [markdown]
# # Map PERSON_ID to Engineer in BoM

# %%
# let's not faff around - just map at a function group level.
# none of the other data is good enough / complete enough to work

engineer_dict = {
    'Body Exterior':'NPETTETT', 
    'Body Structures':'NBATES', 
    'Chassis':'JHEWER', 
    'Electrical':'NHOYLE',
    'Body Interior':'NPETTETT',
    'Styling':'NHOYLE', 
    'Tooling':'NHOYLE', 
    'Development':'NHOYLE', 
    'Powertrain':'DMORRIS'
}


# %%
parts_df['PART_RESPONSIBLE'] = parts_df['FUNCTION_GROUP'].map(engineer_dict)
parts_df['PART_RESPONSIBLE'] = np.where(parts_df['PART_RESPONSIBLE'].isna(), 'NHOYLE', parts_df['PART_RESPONSIBLE'])


# %%
parts_df['PART_LEVEL'] = parts_df['PART_LEVEL'].astype(int)
# parts_df['ENG_PART_REV'] = np.round(parts_df['ENG_PART_REV'], decimals=2)
parts_df['WEIGHT_NET'] = np.round(parts_df['WEIGHT_NET'], decimals = 4)
# parts_df['WEIGHT_NET'] = parts_df['WEIGHT_NET'].truncate(8)


# %%
# temporary fix for zero and decimal QTY
structure_df['QTY'] = np.where(structure_df['QTY'] < 1, 1, structure_df['QTY'])

# %% [markdown]
# # Remove duplicates from parts_df and structure_df
# 
# We've done all the work and calculated the quantities.  IFS only need to be told about each of the parts once, and told of the structures once.
# We don't sum the quantities again - we want 1 steering wheel and reference it in 4 places for the options. Just need to tell IFS once
# 
# If we have this situation, where T50-B0897 and T50-B0020 part and parent are mentioned more than once, we only need to tell IFS once:
# 
# ```   
# SA_Index           Part Number  Parent Part  Issue Level
# 17481.0_T50-B0198  T50-B0198    T50-B0020    3.0            1
# 17481.0_T50-B0365  T50-B0365    T50-B0020    1.0            1
# 17481.0_T50-B0841  T50-B0841    T50-B0020    1.0            1
# 17481.0_T50-B0843  T50-B0843    T50-B0020    1.0            1
# 17481.0_T50-B0845  T50-B0845    T50-B0020    1.0            1
# 17481.0_T50-B0847  T50-B0847    T50-B0020    1.0            1
# >>17481.0_T50-B0897  T50-B0897    T50-B0020    1.0            1<<
# 17481.0_TFF-SA800  TFF-SA800    T50-B0020    1.0            1
# 17501.0_T50-B0198  T50-B0198    T50-B0020    3.0            1
# 17501.0_T50-B0365  T50-B0365    T50-B0020    1.0            1
# 17501.0_T50-B0841  T50-B0841    T50-B0020    1.0            1
# 17501.0_T50-B0843  T50-B0843    T50-B0020    1.0            1
# 17501.0_T50-B0845  T50-B0845    T50-B0020    1.0            1
# 17501.0_T50-B0847  T50-B0847    T50-B0020    1.0            1
# >>17501.0_T50-B0897  T50-B0897    T50-B0020    1.0            1<<
# 17501.0_TFF-AA059  TFF-AA059    T50-B0020    1.0            1
# ```

# %%
parts_df = parts_df.drop_duplicates(subset=['PART_NO','ENG_PART_REV'])
structure_df = structure_df.drop_duplicates()

# %%
# these structures are still duplicated
structure_df[structure_df.duplicated(subset=['PART_NO','SUB_PART_NO'], keep=False)].sort_values(by=['PART_NO','SUB_PART_NO'])
# structure_df[(structure_df['PART_NO'] == 'T50-A5285') & (structure_df['SUB_PART_NO'] == 'TFF-SA907')]

# %% [markdown]
# # Increment ENG_PART_REV
# 
# If there is a released part already in IFS with the same ENG_PART_REV, and we're changing the structure of the part, IFS will not like it.  We need to increment this value to prevent there being issues in IFS migration.  

# %%
# to stop trailing zeros when writing out to csv
parts_df['ENG_PART_REV'] = parts_df['ENG_PART_REV'].apply(str)
structure_df['PART_REV'] = structure_df['PART_REV'].apply(str)
structure_df['SUB_PART_REV'] = structure_df['SUB_PART_REV'].apply(str)

# %% [markdown]
# # Build a DELTA

# %%
# /**/ makes this recursive through folders in the project specfied
from glob import glob

try:
    glob(outdir + '/**/Structure_*' + prev_timestamps + '.txt', recursive = True)[0]
    previous_struct_file = glob(outdir + '/**/Structure_*' + prev_timestamps + '.txt', recursive = True)[0]
    previous_part_file = glob(outdir + '/**/Part*' + prev_timestamps + '.txt', recursive = True)[0]
except:
    print('No previous files found for timestamp {}'.format(prev_timestamps))
    print("location: {}".format(outdir + '/**/Part*' + prev_timestamps + '.txt'))
    exit()

# use dtype in read_csv to capture trailing zeros in ENG_PART_REV
path = os.path.join(base, 'IFS', previous_part_file)
try:
    with open(path, "rb") as f:
        prev_part = pd.read_csv(f, sep='\t', dtype={'ENG_PART_REV':str}) 
        # sheetnames = [sheet for sheet in f.sheet_names]
except (FileNotFoundError):
    logit.critical("File not found: {}".format(path))

try:
    path = os.path.join(base, 'IFS', previous_struct_file)
    with open(path, "rb") as f:
        prev_struct = pd.read_csv(f, sep='\t', dtype={'SUB_PART_REV':str}) 
        # sheetnames = [sheet for sheet in f.sheet_names]
except (FileNotFoundError):
    logit.critical("File not found: {}".format(path))

# %%
previous_part_file

# %%
# derive incrementer from previous part file
# incrementer = prev_part['ENG_PART_REV'].str.split('.').str[-1].astype(int).unique()
incrementer = prev_part['ENG_PART_REV'].astype(str).str.split('.').str[-1].astype(int).unique()
try:
    len(incrementer) == 1
except:
    logit.exception("More than one migration revision found in previous file")
    raise

incrementer = incrementer[0]
incrementer += 1


# %%
# there is a limit of 6 chars on length of PART_REV fields.  We need to drop the minor revisioning for T50 as it's not being used anyway
if project == 'T50':
    parts_df['ENG_PART_REV'] = parts_df['ENG_PART_REV'].str.split('.').str[0]
    structure_df['PART_REV'] = structure_df['PART_REV'].str.split('.').str[0]
    structure_df['SUB_PART_REV'] = structure_df['SUB_PART_REV'].str.split('.').str[0]

# %%
# this is the amount we'll add to the ENG_PART_REV to avoid issues in IFS
# make a string version of the incrementor
str_incr = '.' + str(incrementer)
parts_df['ENG_PART_REV'] = parts_df['ENG_PART_REV'] + str_incr
structure_df['PART_REV'] = structure_df['PART_REV'] + str_incr
structure_df['SUB_PART_REV'] = structure_df['SUB_PART_REV'] + str_incr

# %%
# build a dictionary from a copy
dict_part_compare = {1:prev_part.copy(),2:parts_df.copy()}
dict_struct_compare = {1:prev_struct.copy(),2:structure_df.copy()}

if project == 'T50':
    # before comparision we need to get rid of the end migration revision as we know it will be different.
    # We can do this by converting to int and dropping everything after the decimal place
    # dict_part_compare[1]['ENG_PART_REV'] = dict_part_compare[1]['ENG_PART_REV'].str[0]
    # dict_part_compare[2]['ENG_PART_REV'] = dict_part_compare[2]['ENG_PART_REV'].str[0]
    # dict_struct_compare[1]['PART_REV'] = dict_struct_compare[1]['PART_REV'].str[0]
    # dict_struct_compare[2]['PART_REV'] = dict_struct_compare[2]['PART_REV'].str[0]
    # dict_struct_compare[1]['SUB_PART_REV'] = dict_struct_compare[1]['SUB_PART_REV'].str[0]
    # dict_struct_compare[2]['SUB_PART_REV'] = dict_struct_compare[2]['SUB_PART_REV'].str[0]
    dict_part_compare[1]['ENG_PART_REV'] = dict_part_compare[1]['ENG_PART_REV'].astype(str).str.split('.').str[0]
    dict_part_compare[2]['ENG_PART_REV'] = dict_part_compare[2]['ENG_PART_REV'].astype(str).str.split('.').str[0]
    dict_struct_compare[1]['PART_REV'] = dict_struct_compare[1]['PART_REV'].astype(str).str.split('.').str[0]
    dict_struct_compare[2]['PART_REV'] = dict_struct_compare[2]['PART_REV'].astype(str).str.split('.').str[0]
    dict_struct_compare[1]['SUB_PART_REV'] = dict_struct_compare[1]['SUB_PART_REV'].astype(str).str.split('.').str[0]
    dict_struct_compare[2]['SUB_PART_REV'] = dict_struct_compare[2]['SUB_PART_REV'].astype(str).str.split('.').str[0]    

else:
    # before comparision we need to get rid of the end migration revision as we know it will be different
    dict_part_compare[1]['ENG_PART_REV'] = dict_part_compare[1]['ENG_PART_REV'].str[:3]
    dict_part_compare[2]['ENG_PART_REV'] = dict_part_compare[2]['ENG_PART_REV'].str[:3]
    dict_struct_compare[1]['PART_REV'] = dict_struct_compare[1]['PART_REV'].str[:3]
    dict_struct_compare[2]['PART_REV'] = dict_struct_compare[2]['PART_REV'].str[:3]
    dict_struct_compare[1]['SUB_PART_REV'] = dict_struct_compare[1]['SUB_PART_REV'].str[:3]
    dict_struct_compare[2]['SUB_PART_REV'] = dict_struct_compare[2]['SUB_PART_REV'].str[:3]

dict_struct_compare2=pd.concat(dict_struct_compare)
dict_part_compare2=pd.concat(dict_part_compare)

# ignore_cols = ['WEIGHT_NET']
dict_part_compare2.WEIGHT_NET = np.round(dict_part_compare2.WEIGHT_NET,4).astype(str)

subset_cols = []
if ignore_cols_for_comparison is not None:
    print ("cols to ignore {}".format(ignore_cols_for_comparison))
    subset_cols = dict_part_compare2.drop(columns=ignore_cols_for_comparison).columns
    # delta_parts and delta_struct have the rows with changes
    delta_parts = dict_part_compare2[dict_part_compare2['PART_LEVEL']>=5].drop_duplicates(subset=subset_cols, keep=False)

else:
    print ("no cols to ignore")
    delta_parts = dict_part_compare2[dict_part_compare2['PART_LEVEL']>=5].drop_duplicates(keep=False)

delta_struct = dict_struct_compare2.drop_duplicates(keep=False)


# %%
# get the PART_NO of any individual PARTs that have changed and the SUB_PART_NO of any parent/assembly that's changed
# changed_parts = set(delta_parts['PART_NO'].tolist() + delta_struct['PART_NO'].tolist())
changed_parts = set(delta_parts['PART_NO'].tolist())

# %%
# get the assembly (SA_Index) for any assembly that has those Part numbers
delta_sa_index = cleansed_df['SA_Index'].str.split('_').str[0][cleansed_df['Part Number'].isin(changed_parts)].tolist()

# %%
# set to give us a unique list
sa_set = set(delta_sa_index)

delta_df = pd.DataFrame()
rel_delta_df = pd.DataFrame()
unrel_delta_df = pd.DataFrame()

# create regex pattern for word match at start of string followed by any number of chars
if len(sa_set) > 0:
    pat = '|'.join(r"\b^{}.*\b".format(x) for x in sa_set)
    # build the delta_df
    delta_df = cleansed_df[cleansed_df['SA_Index'].str.contains(pat)]
    # sort it
    delta_df = delta_df.sort_values('orig_sort')
else:
    logit.warning("No changes found for this delta")
    print ("No changes found for this delta")
    


# %% [markdown]
# # Check for unreleased parts/assemblies

# %%

# find unreleased parts/assemblies.  We can't pass any assembly to IFS that isn't completely released.
# must check there is something in the sa_set
if len(sa_set) > 0:
    unrel_sa_set = set(delta_df['SA_Index'][delta_df['Release Status'] != 'REL'].str.split('_').str[0])

# get the fully released assemblies by ignoring the ones containing unreleased sa_index
# must check there is something in the unrel_sa_set
if len(unrel_sa_set) > 0:
    pat = '|'.join(r"\b^{}.*\b".format(x) for x in unrel_sa_set)
    rel_delta_df = delta_df[~delta_df['SA_Index'].str.contains(pat)]
    # get the unrel rows for writing out the warning messages
    unrel_delta_df = delta_df[delta_df['SA_Index'].str.contains(pat)]
else:
    # there are no non released sa to worry about.  This will return an empty df
    logit.info("There are no unreleased parts or assemblies to worry about")
    rel_delta_df = delta_df


# find blank source codes and drop the whole assembly
empty_sc = set(rel_delta_df['SA_Index'][rel_delta_df['Source Code'].isna()].str.split('_').str[0])

# remove any assemblies with empty source codes
if len(empty_sc) > 0:
    pat = '|'.join(r"\b^{}.*\b".format(x) for x in empty_sc)
    rel_sc_delta_df = rel_delta_df[~rel_delta_df['SA_Index'].str.contains(pat)]

else:
    logit.info("There are no blank source codes to worry about")
    rel_sc_delta_df = rel_delta_df



# %% [markdown]
# # Check for Make without Buys
# 
# IFS won't handle parents of Make where there are no child parts to buy
# 
# if MAKE and there is no BUY below next row's part level less than or equal to current part level we have a MAKE without a BUY

# %%
def check_make_no_buy(df):
    # if MAKE and there is no BUY below next row's part level less than or equal to current part level we have a MAKE without a BUY
    # df['PROVIDE'] = np.where(df['Source Code'].isin(['AIH','MIH','MOB']),'Make','Buy')
    make_no_buy = list(df[(df['Source Code'].isin(['AIH','MIH','MOB'])) & (df['Part Level'].shift(-1) <= df['Part Level'])].SA_Index)
    make_no_buy = sorted(make_no_buy)
    return make_no_buy

# %%
#i think we can reset the index for delta_df without impacting 
make_no_buy = check_make_no_buy(delta_df)

# remove any assemblies with make and no buy
make_no_buy_df = pd.DataFrame()
if len(make_no_buy) > 0:
    pat = '|'.join(r"\b^{}.*\b".format(x.split('_')[0]) for x.split('_')[0] in make_no_buy)
    rel_sc_delta_df = rel_sc_delta_df[~rel_sc_delta_df['SA_Index'].str.contains(pat)]

    # get the make no buy part for writing out the warning messages
    make_no_buy_df = delta_df[delta_df['SA_Index'] == x]

# %%
# pass all the remaining parts left in the 
delta_parts_for_all_sa = rel_sc_delta_df['Part Number'].unique().tolist()


# %%
logit.info("{} Assemblies have changed".format(len(sa_set)))
logit.info("{} changed assemblies are not completely released so won't be processed".format(len(unrel_sa_set)))
logit.warning("{} blank source codes that will stop the whole assembly being released".format(len(empty_sc)))
logit.warning("{} Make parts without any Buy child parts that will stop the whole assembly being migrated".format(len(make_no_buy)))
if unrel_delta_df.shape[0] > 0:
    for i, row in unrel_delta_df[['Function Group', 'Sub Group', 'Part Number', 'Parent Part', 'Release Status', 'SA_Index']].iterrows():
        logit.warning("Part of Unreleased assembly and not processed: {} {} {} {} {}".format(row['Function Group'], row['Sub Group'], row['Part Number'], row['Parent Part'], row['Release Status']))

if make_no_buy_df.shape[0] > 0:
    for i, row in make_no_buy_df[['Function Group', 'Sub Group', 'Part Number', 'Parent Part', 'Source Code', 'SA_Index']].iterrows():
        logit.warning("Make Part with no buy child so assembly not processed: {} {} {} {} {}".format(row['Function Group'], row['Sub Group'], row['Part Number'], row['Parent Part'], row['Source Code']))
        
print ("{} Assemblies have changed".format(len(sa_set)))
print ("{} changed assemblies that are not completely released so won't be migrated".format(len(unrel_sa_set)))
print ("{} blank source codes".format(len(empty_sc)))
print ("{} Make parts without any Buy child parts".format(len(make_no_buy)))


# %%
# delta_all_parts = cleansed_df['Part Number'][cleansed_df['SA_Index'].str.contains('^{}'.format(sa_set))].to_list()
# delta_parts_df = parts_df[parts_df['PART_NO'].isin(delta_parts_for_all_sa)]
delta_structure_df = structure_df[structure_df['PART_NO'].isin(delta_parts_for_all_sa)]

delta_all_parts = set(delta_parts_for_all_sa + delta_structure_df['SUB_PART_NO'].tolist())

delta_parts_df = parts_df[parts_df['PART_NO'].isin(delta_all_parts)]


# %%
def highlight_diff(data, color='pink'):
    # Define html attribute
    attr = 'background-color: {}'.format(color)
    other = data.xs('Previous', axis='columns', level=-1)
    # Where data != other set attribute
    return pd.DataFrame(np.where((data.ne(other, level=0)), attr, ''),
                        index=data.index, columns=data.columns)    

# %%
def compare_parts(df):
    try:
        df_all = pd.concat([df.loc[1,].set_index('PART_NO'), df.loc[2,].set_index('PART_NO')], axis='columns', keys=['Previous','Current'])
        df_final = df_all.swaplevel(axis='columns')[delta_parts.columns[1:]].fillna('')
    except:
        df_all = pd.concat([pd.DataFrame('', columns=delta_parts.columns, index=delta_parts['PART_NO']), delta_parts.loc[2,].set_index('PART_NO')], axis='columns', keys=['Previous','Current'])
        df_final = df_all.swaplevel(axis='columns')[delta_parts.columns[1:]].fillna('')

    return df_final

# %%
def compare_struct(df):
    indx = ['PART_NO','SUB_PART_NO']
    df_all = pd.concat([df.loc[1,].set_index(indx), df.loc[2,].set_index(indx)], axis='columns', keys=['Previous','Current'])
    df_final = df_all.swaplevel(axis='columns')[df.drop(columns=indx).columns].fillna('')

    return df_final

# %% [markdown]
# # Build a test file
# 
# Provide a part number and will build test files

# %%
# test_part = 'T50-A8115'
# # filter for test_part in PART_NO in structure file.  Store SUB_PART_NOs

# # test_parts = structure_df['SUB_PART_NO'][structure_df['PART_NO'] == 'T50-A8423'].to_list()

# # use the sa_index to find whole assembly
# SA_Index = cleansed_df['SA_Index'].str.split('_').str[0][cleansed_df['Part Number'] == test_part]
# test_parts = cleansed_df['Part Number'][cleansed_df['SA_Index'].str.contains('^{}'.format(SA_Index))].to_list()
# # add the test part to the list of SUB_PART_NOs
# test_parts.append(test_part)
# # go back to structure file and look for list of parts in PART_NO
# test_structure_df = structure_df[structure_df['PART_NO'].isin(test_parts)]

# all_test_parts = set(test_parts + test_structure_df['SUB_PART_NO'].tolist())

# test_parts_df = parts_df[parts_df['PART_NO'].isin(all_test_parts)]

# %%
cleansed_df

# %%
# Validation checks before writing out.  Don't write files without an error_count of zero

TEST=False
# delta being set / overridden at the top of the scrpt with args.  Default is DELTA = True
# DELTA=True

if TEST:
    print ("*** TEST MODE ***")
    out_structure_df = test_structure_df
    out_parts_df = test_parts_df
else:
    out_structure_df = structure_df
    out_parts_df = parts_df
     

error_count = 0

# check for zero quantities
zero_quantities = out_structure_df[out_structure_df['QTY'] == 0]
if zero_quantities.shape[0] > 0:
    zero_quantities.to_excel(os.path.join(base, '{}_Zero_Quantity.xlsx'.format(project)))
    logit.error("Zero Quantities found - needs resolving first!")
    print ("Zero Quantities found - needs resolving first!")
    error_count += 1

# check for decimal quantities
decimal_quantities = out_structure_df[(out_structure_df['QTY'] > 0) & (out_structure_df['QTY'] < 1)]
if decimal_quantities.shape[0] > 0:
    decimal_quantities.to_excel(os.path.join(base, '{}_Decimal_Quantity.xlsx'.format(project)))
    logit.error("Decimal Quantities found - needs resolving first!")
    print ("Decimal Quantities found - needs resolving first!")
    error_count += 1

# check all parts present
orphaned_parts = []
orphaned_parts = pd.merge(out_parts_df, out_structure_df, left_on='PART_NO', right_on='SUB_PART_NO', how='left', indicator=True)
orphaned_parts = orphaned_parts[['PART_NO_x','SUB_PART_NO']][orphaned_parts['_merge'] == 'left_only']
if orphaned_parts.shape[0] > 1:
    logit.error("Expecting just the top CAR part to not have any parent")
    logit.info("orphaned parts {}".format(orphaned_parts))
    print ("Expecting just the top CAR part to not have any parent")
    print (orphaned_parts)
    print ("")
    error_count =+ 1

# check all master parts have child parts
no_child_part = []
no_child_part = pd.merge(out_structure_df, out_parts_df, left_on='SUB_PART_NO', right_on='PART_NO', how='left', indicator=True)
no_child_part = no_child_part[['PART_NO_x','SUB_PART_NO']][no_child_part['_merge'] == 'left_only']
if no_child_part.shape[0] > 0:
    logit.error ("Not expecting any parts without sub parts")
    logit.error (no_child_part)
    print ("Not expecting any parts without sub parts")
    print (no_child_part)
    error_count =+ 1

# st
sub_part_rev_check = []
sub_part_rev_check = pd.merge(out_structure_df, out_parts_df, left_on=['SUB_PART_NO','SUB_PART_REV'], right_on=['PART_NO','ENG_PART_REV'], indicator=True, how='left')
sub_part_rev_check = sub_part_rev_check[sub_part_rev_check['_merge'] == 'left_only']
if sub_part_rev_check.shape[0] > 0:
    logit.error ("SUB_PART_REV and ENG_PART_REV do not match")
    logit.error (sub_part_rev_check)
    print ("SUB_PART_REV and ENG_PART_REV do not match")
    print (sub_part_rev_check)
    error_count =+ 1        

# check sub part rev matches in parts and structure files
part_rev_check = out_structure_df[out_structure_df.PART_REV.isna()]
if part_rev_check.shape[0] > 0:
    logit.error ("Can't have a blank PART_REV for these parts in structure file")
    logit.error (part_rev_check)
    print ("Can't have a blank PART_REV for these parts in structure file")
    print (part_rev_check)
    error_count =+ 1

# find master parts in parts_df and check PART_REV and SUB_PART_REV match
master_part = out_structure_df[['PART_REV','PART_NO']]
missing_master = pd.merge(master_part, out_structure_df, left_on=['PART_REV','PART_NO'], right_on=['SUB_PART_REV','SUB_PART_NO'], how='left', indicator=True)
missing_master = missing_master[missing_master['_merge'] == 'left_only']
missing_master = missing_master[~missing_master['PART_NO_x'].str.contains('-CAR')]
if TEST:
    # for a test file we won't need the structure for the actual assembly we've specified
    missing_master = missing_master[~missing_master['PART_NO_x'].str.contains(test_part)]

if missing_master.shape[0] > 0:
    logit.error ("missing_master: Must find PART_REV / PART_NO combo in SUB_PART_REV / SUB_PART_NO")
    logit.error (missing_master)
    print ("missing_master: Must find PART_REV / PART_NO combo in SUB_PART_REV / SUB_PART_NO")
    print (missing_master)
    error_count =+ 1   


if error_count == 0: 

    outfile_part = 'Part_{}_{}_{}_{}'.format(project, env, incrementer, timestr)
    outfile_structure = 'Structure_{}_{}_{}_{}'.format(project, env, incrementer, timestr)

    if TEST:
        outfile_part = outfile_part + '_' + test_part
        outfile_structure = outfile_structure + '_' + test_part

    def write_to_excel(df, outfile):
        with pd.ExcelWriter(os.path.join(outdir, outfile), engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name = 'Sheet1', index=False)
                ws = writer.sheets['Sheet1']
                wb = writer.book
                excel_formatting.adjust_col_width_from_col(ws)
    
    def write_to_csv(df, outfile):
        df.to_csv(os.path.join(outdir, outfile),sep='\t', index=False, encoding='utf-8')
        # parts_df.to_csv(os.path.join(outdir, outfile),sep='\t', index=False, encoding='utf-8')

    if DELTA:
        logit.info("*** Writing out changes since {} file ***".format(prev_timestamps))
        print ("*** Writing out changes since {} file ***".format(prev_timestamps))
        write_to_excel(delta_parts_df, 'DELTA_{}.xlsx'.format(outfile_part))
        write_to_csv(delta_parts_df, 'DELTA_{}.txt'.format(outfile_part))
        write_to_excel(delta_structure_df, 'DELTA_{}.xlsx'.format(outfile_structure))
        write_to_csv(delta_structure_df, 'DELTA_{}.txt'.format(outfile_structure))

        # write out the highlighted changes
        compare_file = 'COMPARE_{}_vs_{}.xlsx'.format(prev_timestamps, timestr)
        compare_out = os.path.join(outdir, compare_file)
        df_final_parts = compare_parts(delta_parts)
        try:
            delta_struct.loc[1,]
            try:
                delta_struct.loc[2,]
                df_final_struct = compare_struct(delta_struct)
            except (KeyError):
                print ("No current parts - assumed all changes are parts being removed")
                # nothing to compare so just provide an empty delta_struct
                df_final_struct = delta_struct
        except (KeyError) as e:
            print ("No previous parts - assumed all changes are parts being added")
            df_final_struct = delta_struct
        # df_final_part.style.apply(highlight_diff, axis=None).to_excel(compare_out, engine='openpyxl',)
        with pd.ExcelWriter(compare_out) as writer:
            try:
                df_final_struct.loc[1,]
                try:
                    df_final_struct.loc[2,]
                    # highlight the differences between index 1 and 2
                    df_final_struct.style.apply(highlight_diff, axis=None).to_excel(writer, sheet_name='Structure')
                except (KeyError) as e:
                    # nothing in 2 to compare.  Write out the structures involved with the parts
                    delta_structure_df.to_excel(writer, sheet_name='Structure')
            except (KeyError) as e:
                # nothing in 1 to compare.  Write out the structures involved with the parts
                delta_structure_df.to_excel(writer, sheet_name='Structure')

            df_final_parts.style.apply(highlight_diff, axis=None).to_excel(writer, sheet_name='Parts')

        # write out the delta df file for Lorena to file a bom-like file
        delta_bom_file = 'Changed_Assemblies_BOM_{}_{}_vs_{}.xlsx'.format(project, prev_timestamps, timestr)
        delta_bom_out = os.path.join(outdir, delta_bom_file)
        write_to_excel(delta_df, delta_bom_out)

    write_to_excel(out_parts_df, outfile_part + '.xlsx')
    write_to_csv(out_parts_df, outfile_part + '.txt')
    write_to_excel(out_structure_df, outfile_structure + '.xlsx')
    write_to_csv(out_structure_df, outfile_structure + '.txt')



# %%
logit.info('Completed')

for handler in logit.handlers:
    if isinstance(handler, logging.FileHandler):
        handler.close()

# %%
for handler in logit.handlers:
    if isinstance(handler, logging.FileHandler):
        handler.close()

# %%



